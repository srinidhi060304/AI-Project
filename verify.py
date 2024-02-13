import smtplib
import time
import random
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
import pytesseract
import re
from PIL import Image
import cv2
from queue import Queue
import threading
import subprocess


# Function to get user input with a timeout
def get_user_input(queue):
    global user_input
    user_input = input(f"Enter the OTP sent to your email {email}: ")
    queue.put(True)  # Notify the main thread that the input is received

# Initialize the camera capture
cap = cv2.VideoCapture(0)  # 0 for the default camera, or you can use other indexes for additional cameras if available

if not cap.isOpened():
    print("Error: Cannot open the camera.")
    exit()

capture_duration = 5  # 5 seconds
start_time = time.time()
remaining_time = capture_duration

while remaining_time > 0:
    ret, frame = cap.read()

    if not ret:
        print("Error: Cannot read a frame.")
        break

    elapsed_time = time.time() - start_time

    if elapsed_time >= 1:
        print(f"Countdown: {int(remaining_time)} seconds")
        start_time = time.time()
        remaining_time -= 1

    # Display the frame in a window
    cv2.imshow("Camera Feed", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the camera and close the OpenCV window
cap.release()
cv2.destroyAllWindows()

# Save the captured frame as an image with a specific name (e.g., "captured_image.jpg")
cv2.imwrite("id.jpg", frame)
print("Image saved as 'id.jpg' in the current folder.")



pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"
path=r"D:\ASEB\Semester 3\Projects\AI\ai project start part\Details.xlsx"
file_object=openpyxl.load_workbook(path).active

row = file_object.max_row  
col=file_object.max_column
image = Image.open('id.jpg')
extracted_text = pytesseract.image_to_string(image)

pattern = r'BL\s*[._\-\s]*EN\s*[._\-\s]*([a-zA-Z0-9]{10})'
matches = re.findall(pattern, extracted_text, re.IGNORECASE)

if matches:
    for match in matches:
        # Convert the match to lowercase
        match = match.lower()
else:
    print("No matching strings were found in the text.")

reg="bl.en."+""+match
print(reg)
n=0
flag=0
name=""
branch=""
section=""
email=""
for i in range(3,row+1):
    if(file_object.cell(row=i,column=1).value==reg):
        n=i
        name=file_object.cell(row=i,column=2).value
        branch=file_object.cell(row=i,column=3).value
        section=file_object.cell(row=i,column=4).value
        email=file_object.cell(row=i,column=5).value
        flag=1
        break

if(flag==1):
    print("\n\nRegistration Number: ",reg)
    print("\nName: ",name)
    print("\nBranch: ",branch)
    print("\nSection: ",section)
    print("\nEmail: ",email)
else:
    print("Not Found")
    exit()

otp = str(random.randint(100000, 999999))
sender_email = '50shadesincorp@gmail.com'
sender_password = 'iorh sdtp jdlv ujmb'
receiver_email = email

msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = receiver_email
msg['Subject'] = 'Your OTP Verification Code'

msg.attach(MIMEText(f'Welcome {name},\nYour OTP is: {otp}\nThank You For Ordering\nHave a Great Day', 'plain'))
try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(sender_email, sender_password)
    server.sendmail(sender_email, receiver_email, msg.as_string())
    print("OTP sent successfully!")
except Exception as e:
    print(f"Error sending OTP: {e}")
finally:
    server.quit()

# Add a timer for 2 minutes (120 seconds) for OTP entry
otp_timeout = 120  # 2 minutes
start_time_otp = time.time()

# Use a queue to communicate between threads
input_queue = Queue()

# Start a thread for user input
input_thread = threading.Thread(target=get_user_input, args=(input_queue,))
input_thread.start()

# Wait for either the input or the timeout
input_thread.join(timeout=otp_timeout)
# Check if the input is received
if input_queue.qsize() > 0:
    # Input received before timeout
    if user_input == otp:
        print("OTP verified successfully!")
        script_name=r"D:\ASEB\Semester 3\Projects\AI\ai_agent\ai_agent\AI_customer.py"
        subprocess.call(['python', script_name])
    else:
        print("Incorrect OTP. Please try again.")
else:
    print("\nOTP entry timeout. Please request a new OTP and try again.")

